from __future__ import print_function
# from builtins import str
# from builtins import object
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import os
from .data_utils import make_dir

class XlsxRecorder(object):
    """
    xlsx recorder for results
    including two recorder: one for current experiments, record details of results changed by iteration
    the other is for record the summary of different expreiments, which is saved by  summary_path

    1. detailed results: saved in fig_save_path/results.xlsx
        ** Sheet1: #total_filename x #metrics,   along each row direction, are the records by iteraton
        ** batch_0: #batch_filename x #metric_by_label , along each column direction, are the records by iteration
        ** batch_1: same as batch_0
        ** ......
    2. task results: saved in ../data/summary.xlsx
        ** Sheet1: task_name * #metrics recorded by iteration
    """
    def __init__(self, expr_name, saving_path='', folder_name=''):
        self.expr_name = expr_name
        if not len(saving_path):
            self.saving_path = '../data/'+expr_name  #saving_path
        else:
            self.saving_path = saving_path
        self.saving_path = os.path.abspath(self.saving_path)
        self.folder_name = folder_name
        if len(folder_name):
            self.saving_path = os.path.join(self.saving_path, folder_name)
        """path of saving excel, default is the same as the path of saving figures"""
        self.writer_path = None
        self.xlsx_writer =  None
        self.summary_path = '../data/summary.xlsx'
        """the path for summary, which can record results from different experiments"""
        self.measures = ['iou', 'precision', 'recall', 'dice']
        """measures to record"""
        self.batch_count = {}
        self.row_space = 50
        self.column_space = 10
        self.start_row = 0
        self.summary = None
        self.avg_buffer = {}
        self.iter_info_buffer = []
        self.name_list_buffer = []
        self.init_summary()
        print("the update space in detailed files is {}".format(self.row_space))

    def init_summary(self):
        """ init two recorders, initilzation would create a new recorder for this experiment, recording all details
            at the same time it would load the data from summary recorder, then it would append the new experiment summary to summary recorder
        """
        if not os.path.exists(self.saving_path ):
            os.makedirs(self.saving_path )
        self.writer_path = os.path.join(self.saving_path, 'results.xlsx')
        writer = pd.ExcelWriter(self.writer_path, engine='xlsxwriter')
        df = pd.DataFrame([])
        df.to_excel(writer)
        worksheet = writer.sheets['Sheet1']
        worksheet.set_column(1, 1000, 30)
        writer.save()
        writer.close()
        self.writer_book = load_workbook(self.writer_path)
        self.xlsx_writer = pd.ExcelWriter(self.writer_path, engine='openpyxl')
        self.xlsx_writer.book = self.writer_book
        self.xlsx_writer.sheets = dict((ws.title, ws) for ws in self.writer_book.worksheets)
        if not os.path.exists(self.summary_path):
            writer = pd.ExcelWriter(self.summary_path, engine = 'xlsxwriter')
            df = pd.DataFrame([])
            df.to_excel(writer)
            worksheet = writer.sheets['Sheet1']
            worksheet.set_column(1, 1000, 30)
            writer.save()
            writer.close()


    def set_batch_based_env(self,name_list,batch_id):
        # need to be set before each saving operation
        self.name_list = name_list
        self.sheet_name = 'batch_'+ str(batch_id)
        if self.sheet_name not in self.batch_count:
            self.batch_count[self.sheet_name] = -1
            self.name_list_buffer += self.name_list
        self.batch_count[self.sheet_name] += 1
        count = self.batch_count[self.sheet_name]
        self.start_row = count * self.row_space
        self.start_column = 0


    def set_summary_based_env(self):
        self.sheet_name = 'Sheet1'
        self.start_row = 0

    def put_into_avg_buff(self, result, iter_info):
        """
        # avg_buffer is to save avg_results from each iter, from each batch
        # iter_info: string contains iter info
        # the buffer is organized as { iter_info1: results_list_iter1, iter_info2:results_list_iter2}
        # results_list_iter1 : [batch1_res_iter1, batch2_res_iter1]
        # batch1_res_iter1:{metric1: result, metric2: result}
        """
        if iter_info not in self.avg_buffer:
            self.avg_buffer[iter_info] = []
            self.iter_info_buffer += [iter_info]
        self.avg_buffer[iter_info] += [result]

    def merge_from_avg_buff(self):
        """
        # iter_info: string contains iter info
        # the buffer is organized as { iter_info1: results_list_iter1, iter_info2:results_list_iter2}
        # results_list_iter1 : [batch1_res_iter1, batch2_res_iter1]
        # batch1_res_iter1:{metric1: result, metric2: result}
        # return: dic: {iter_info1:{ metric1: nFile x 1 , metric2:...}, iter_info2:....}
        """
        metric_avg_dic={}
        for iter_info,avg_list in list(self.avg_buffer.items()):
            metric_results_tmp = {metric: [result[metric] for result in avg_list] for metric in
                                  self.measures}
            metric_avg_dic[iter_info] = {metric: np.concatenate(metric_results_tmp[metric], 0) for metric in metric_results_tmp}
        return  metric_avg_dic



    def saving_results(self,sched, results=None, info=None, averaged_results=None):
        """
        the input results should be different for each sched
        batch: the input result should be dic , each measure inside should be B x N_label
        buffer: the input result should be dic, each measure inside should be N_img x 1
        summary: no input needed, the summary could be got from the buffer

        :param results:
        :param sched:
        :param info:
        :return:
        """
        if sched == 'batch':
            label_info = info['label_info']
            iter_info = info['iter_info']
            self.saving_all_details(results,averaged_results,label_info,iter_info)

        elif sched == 'buffer':
            iter_info = info['iter_info']
            self.put_into_avg_buff(results,iter_info)
        elif sched == 'summary':
            self.summary_book = load_workbook(self.summary_path)
            self.summary_writer = pd.ExcelWriter(self.summary_path,engine='openpyxl')
            self.set_summary_based_env()
            metric_avg_dic = self.merge_from_avg_buff()
            self.saving_label_averaged_results(metric_avg_dic)
            self.saving_summary(metric_avg_dic)
            self.save_figs_for_batch(metric_avg_dic)
            self.xlsx_writer.close()
            self.summary_writer.close()
        else:
            raise ValueError("saving method not implemented")


    def saving_label_averaged_results(self, results):
        """
        # saved by iteration
        # results: dic: {iter_info1:{ metric1: nFile x 1 , metric2:...}, iter_info2:....}
        # saving the n_File*nAvgMetrics into xlsx_writer
        # including the iter_info
        """
        start_column = 0
        results_summary = {iter_info: {metric:np.mean(results[iter_info][metric]).reshape(1,1) for metric in self.measures} for iter_info in self.iter_info_buffer}
        for iter_info in self.iter_info_buffer:
            iter_expand = {metric: np.squeeze(np.concatenate((results[iter_info][metric], results_summary[iter_info][metric]), 0)) for metric in self.measures}
            df = pd.DataFrame.from_dict(iter_expand)
            df = df[self.measures]
            try:
                df.index = pd.Index(self.name_list_buffer+['average'])
            except:
                print("DEBUGGING !!, the iter_expand is {},\n  self.name_list_buffer is {},\n results summary is {} \n results{}\n".format(iter_expand,self.name_list_buffer, results_summary,results))
            df.to_excel(self.xlsx_writer, sheet_name=self.sheet_name, startcol=start_column, index_label=iter_info)
            start_column += self.column_space
        self.xlsx_writer.save()





    def saving_summary(self, results):
        """
        # saved by iteration
        # saving the 1*nAvgMetrics into summary_book_path
        # including the task name and iter_info
        """
        self.summary_writer.book = self.summary_book
        self.summary_writer.sheets = dict((ws.title, ws) for ws in self.summary_book.worksheets)
        col_name_list = [metric+iter_info for metric in self.measures for iter_info in self.iter_info_buffer]
        results_summary = {metric+iter_info: np.mean(results[iter_info][metric]).reshape(1) for metric in self.measures for iter_info in self.iter_info_buffer}
        df = pd.DataFrame.from_dict(results_summary)
        df = df[col_name_list]
        df.index = pd.Index([self.expr_name+self.folder_name])
        startrow = self.summary_writer.sheets['Sheet1']._current_row + 1
        df.to_excel(self.summary_writer, startrow = startrow)
        self.summary_writer.save()







    def saving_all_details(self, results,averaged_results,label_info,iter_info):
        # saved by batch_list x  n_metric_*len_label_list

        label_list = label_info
        data = {measure: np.concatenate((results[measure], averaged_results[measure]),0) for measure in self.measures}
        row_index = np.asarray(self.name_list+['average'])
        column_index = [measure + '_' + str(label) for measure in self.measures for label in label_list]
        formated_data = {measure + '_' + str(label): data[measure][:, j] for measure in self.measures for j, label in
                         enumerate(label_list)}
        df = pd.DataFrame.from_dict(formated_data)
        df.index = pd.Index(row_index)
        df = df[column_index]
        try:
            df.to_excel(self.xlsx_writer, sheet_name=self.sheet_name, startrow= self.start_row, index_label= iter_info)
        except:
            df.to_excel(self.xlsx_writer, sheet_name=self.sheet_name, startrow=self.start_row, index_label=iter_info)
        worksheet = self.xlsx_writer.sheets[self.sheet_name]
        #worksheet.set_column(1,1000, 30)
        self.xlsx_writer.save()




    def save_figs_for_batch(self,results_summary):
        """
        the function will plot the dice score for each batch and
        :return:
        """
        from matplotlib import pyplot as plt
        import numpy as np

        plt.style.use('fivethirtyeight')
        img_num = len(self.name_list_buffer)
        iter_num = len(self.iter_info_buffer)
        img_dice_np = np.zeros([img_num,iter_num])-1
        x_axis = list(range(iter_num))

        for j,iter_info in enumerate(self.iter_info_buffer):
            for i in range(img_num):
                img_dice_np[i,j] = results_summary[iter_info]['dice'][i][0]

        fig, ax = plt.subplots()
        for i in range(img_num):
            ax.plot(x_axis, img_dice_np[i,:],label=self.name_list_buffer[i])

        ax.legend( prop={'size': 4})
        ax.set( ylabel='average_dice',
               title='dice score for each sample')
        plt.xticks(x_axis)
        ax.set_xticklabels(self.iter_info_buffer,fontdict={'fontsize':5})
        for tick in ax.get_xticklabels():
            tick.set_rotation(55)
        path  = os.path.join(self.saving_path, 'images_dice_iter.png')
        plt.savefig(path,dpi=300, bbox_inches='tight')