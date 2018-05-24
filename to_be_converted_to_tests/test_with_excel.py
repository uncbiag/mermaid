##############################################################################
#
# An example of positioning dataframes in a worksheet using Pandas and
# XlsxWriter.
#
# Copyright 2013-2017, John McNamara, jmcnamara@cpan.org
#

from builtins import str
import pandas as pd
from openpyxl import load_workbook
import numpy as np

# Create some Pandas dataframes from some data.
df1 = pd.DataFrame({'Data': [11, 12, 13, 14,18]})
df2 = pd.DataFrame({'Data': [21, 22, 23, 24]})
df3 = pd.DataFrame({'Data': [31, 32, 33, 34]})
df4 = pd.DataFrame({'Data': [41, 42, 43, 44]})

# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter('pandas_positioning.xlsx', engine='xlsxwriter')

# Position the dataframes in the worksheet.
df1.to_excel(writer, sheet_name='Sheet1')  # Default position, cell A1.
df2.to_excel(writer, sheet_name='Sheet1', startcol=3)
df3.to_excel(writer, sheet_name='Sheet1', startrow=6)

# It is also possible to write the dataframe without the header and index.
df4.to_excel(writer, sheet_name='Sheet1',
             startrow=7, startcol=4, index_label='test', header=False, index=True)

# Close the Pandas Excel writer and output the Excel file.
worksheet = writer.sheets['Sheet1']
#number_of_rows = worksheet.nrows
worksheet.set_column(1,1000, 30)


writer.save()




# col: measure  row: file
book = load_workbook('pandas_positioning.xlsx')
writer = pd.ExcelWriter('pandas_positioning.xlsx',engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
labels =['a','b','c','d','e']
measures=['iou','precision','recall','dice']
names = ['file1', 'file2', 'file3']
#data = {'measures':[measure*len(label) for measure in measures]}
res = np.random.randn(3,5)
data = {'iou':res,'precision':res,'recall':res,'dice':res}
row_index = np.asarray(names)
column_index =  [measure+'_' + str(label) for measure in measures for label in labels ]
formated_data = {measure+'_' + str(label): data[measure][:,j] for measure in measures for j, label in enumerate(labels) }
df = pd.DataFrame.from_dict(formated_data)
df.index = pd.Index(row_index)
df = df[column_index]

df.to_excel(writer, sheet_name ='sheet2')


writer.save()
df.to_excel(writer, sheet_name ='sheet3')
writer.save()


# col: file row: measure
book = load_workbook('pandas_positioning.xlsx')
writer = pd.ExcelWriter('pandas_positioning.xlsx',engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
labels =['a','b','c','d','e']
measures=['iou','precision','recall','dice']
names = ['file1', 'file2', 'file3']
#data = {'measures':[measure*len(label) for measure in measures]}
data = {'iou':res,'precision':res,'recall':res,'dice':res}

column_index = np.asarray(names)
row_index =  [measure+'_' + str(label) for measure in measures for label in labels ]
formated_data = {name: np.concatenate([data[measure][id] for measure in measures]).tolist() for id, name in enumerate(names) }
df = pd.DataFrame.from_dict(formated_data)
df.index = pd.Index(row_index)
df = df[column_index]

df.to_excel(writer, sheet_name ='sheet3')

writer.save()